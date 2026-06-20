import streamlit as st
import pandas as pd
import io
import re
import streamlit.components.v1 as components
from streamlit_gsheets import GSheetsConnection

# ─────────────────────────────────────────────
# IMPORTAÇÕES PARA ESTILIZAÇÃO DO EXCEL
# ─────────────────────────────────────────────
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ─────────────────────────────────────────────
# CONFIGURAÇÃO DA PÁGINA
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Gestão de Pedidos - Padaria e Confeitaria",
    page_icon="🥐",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────
# INICIALIZAÇÃO DE VARIÁVEIS DE SESSÃO
# ─────────────────────────────────────────────
if 'reset_counter_padconf' not in st.session_state:
    st.session_state['reset_counter_padconf'] = 0

if 'usuario_logado_padconf' not in st.session_state:
    st.session_state['usuario_logado_padconf'] = None

# ─────────────────────────────────────────────
# CSS GLOBAL E DE IMPRESSÃO
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@400;500;700&display=swap');

:root {
    --bg-main:        #0d1117;
    --bg-card:        #161b22;
    --bg-sidebar:     #0d1117;
    --brown-dark:     #4A2311;
    --brown-mid:      #D35400;
    --brown-accent:   #E67E22;
    --brown-bright:   #FDEBD0;
    --brown-glow:     rgba(230, 126, 34, 0.25);
    --text-primary:   #e6edf3;
    --text-muted:     #7d8590;
    --text-header:    #FDEBD0; 
    --border:         #21262d;
    --border-active:  #E67E22;
    --row-hover:      rgba(230, 126, 34, 0.08);
    --row-selected:   rgba(230, 126, 34, 0.18);
}

.stApp, .main { background-color: var(--bg-main) !important; color: var(--text-primary) !important; }
html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif !important; }
section[data-testid="stSidebar"] { background-color: var(--bg-sidebar) !important; border-right: 1px solid var(--border); }
section[data-testid="stSidebar"] * { color: var(--text-primary) !important; }
section[data-testid="stSidebar"] .stRadio label { font-size: 14px; }

.stButton > button[kind="primary"], .stFormSubmitButton > button {
    background: linear-gradient(135deg, var(--brown-mid) 0%, var(--brown-accent) 100%) !important;
    color: #fff !important;
    border: 1px solid var(--brown-accent) !important;
    border-radius: 8px !important;
    font-weight: 700 !important;
    letter-spacing: .3px;
    transition: all .2s ease !important;
}
.stButton > button[kind="primary"]:hover, .stFormSubmitButton > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 4px 18px var(--brown-glow) !important;
}
.stButton > button {
    background: var(--bg-card) !important;
    color: var(--text-primary) !important;
    border: 1px solid var(--border) !important;
    border-radius: 8px !important;
    transition: all .2s ease !important;
}
.stButton > button:hover {
    border-color: var(--brown-accent) !important;
    color: var(--brown-bright) !important;
    transform: translateY(-1px) !important;
}
.stTextInput input, .stSelectbox > div > div, .stTextArea textarea {
    background-color: var(--bg-card) !important;
    border: 1px solid var(--border) !important;
    border-radius: 8px !important;
    color: var(--text-primary) !important;
}
.stTextInput input:focus, .stSelectbox > div > div:focus-within, .stTextArea textarea:focus {
    border-color: var(--brown-accent) !important;
    box-shadow: 0 0 0 3px var(--brown-glow) !important;
}
.title-input input {
    font-weight: 700 !important;
    font-size: 16px !important;
    color: var(--brown-bright) !important;
    padding: 2px 8px !important;
    background: transparent !important;
    border: 1px dashed #21262d !important;
}
.title-input input:focus { border: 1px dashed var(--brown-accent) !important; }

[data-testid="stDataEditor"] [data-testid="glideDataEditor"] .gdg-header-cell,
[data-testid="stDataEditor"] .dvn-stack .gdg-header {
    background-color: var(--brown-dark) !important;
    color: var(--text-header) !important;
}

[data-testid="stDataEditor"] {
    border-radius: 10px !important;
    overflow: hidden;
    border: 1px solid var(--brown-mid) !important;
    box-shadow: 0 4px 20px rgba(0,0,0,.4);
    font-size: 12px !important;
}
[data-testid="stDataEditor"] .gdg-cell.gdg-selected,
[data-testid="stDataEditor"] .gdg-cell[data-state="focused"],
[data-testid="stDataEditor"] .gdg-cell[aria-selected="true"] {
    background-color: var(--row-selected) !important;
    outline: 2px solid var(--brown-accent) !important;
    outline-offset: -2px;
}
[data-testid="stDataEditor"] .gdg-row:hover .gdg-cell { background-color: var(--row-hover) !important; }

div[data-testid="stVerticalBlockBorderWrapper"] {
    background-color: var(--bg-card) !important;
    border: 1px solid var(--border) !important;
    border-radius: 12px !important;
    transition: box-shadow .25s ease, border-color .25s ease;
}
div[data-testid="stVerticalBlockBorderWrapper"]:hover {
    border-color: var(--brown-mid) !important;
    box-shadow: 0 6px 24px rgba(0,0,0,.35) !important;
}
[data-testid="stMetric"] {
    background-color: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 10px 10px;
}
[data-testid="stMetricValue"] { color: var(--brown-bright) !important; font-weight: 700; font-size: 1.8rem !important; }
[data-testid="stMetricLabel"] { color: var(--text-muted) !important; }

.topbar-loja {
    background: linear-gradient(90deg, var(--brown-dark) 0%, #1c140d 100%);
    border: 1px solid var(--brown-mid);
    border-radius: 10px;
    padding: 10px 18px;
    margin-bottom: 18px;
    display: flex;
    align-items: center;
    justify-content: space-between;
}
.topbar-left { display: flex; align-items: center; gap: 12px; }
.topbar-title { font-size: 18px; font-weight: 700; color: var(--text-header); }
.topbar-sub { font-size: 11px; color: var(--text-muted); margin-top: 2px; }
.erp-badge { background-color: #2ea043; color: white; padding: 2px 8px; border-radius: 12px; font-size: 10px; font-weight: 600; margin-left: 8px;}

/* IMPRESSÃO */
@media print {
    @page { size: landscape; margin: 5mm; }

    html, body, .stApp, #root, [data-testid="stAppViewContainer"], [data-testid="stMainBlockContainer"], .main, .block-container {
        padding-top: 0 !important;
        padding-bottom: 0 !important;
        margin-top: 0 !important;
        background-color: #ffffff !important;
        background-image: none !important;
        color: #000000 !important;
        width: 100% !important;
        max-width: 100% !important;
    }

    header, [data-testid="stSidebar"], [data-testid="stHeader"] { display: none !important; }
    [data-testid="stElementContainer"],
    [data-testid="stHorizontalBlock"],
    div[data-testid="stVerticalBlockBorderWrapper"] { display: none !important; }
    
    [data-testid="stElementContainer"]:has(#print-section) { display: block !important; width: 100% !important; padding: 0 !important; margin: 0 !important; }
    
    #print-section { display: block !important; width: 100% !important; margin-top: 0 !important; padding-top: 0 !important; }
    #print-section h2 { font-size: 14px !important; margin: 0 0 6px 0 !important; padding-bottom: 3px !important; border-bottom: 2px solid #000 !important; color: #000 !important; display: block !important; text-align: center !important; }
    #print-section h3 { font-size: 11px !important; font-weight: 700 !important; border-bottom: none !important; margin-top: 10px !important; margin-bottom: 3px !important; color: #000 !important; }
    .print-container { width: 100% !important; display: block !important; }

    table.print-table { width: 100% !important; border-collapse: collapse !important; color: #000000 !important; font-family: 'IBM Plex Sans', sans-serif !important; line-height: 1.2 !important; display: table !important; table-layout: fixed !important; margin-bottom: 5px !important; }
    table.print-table th, table.print-table td { border: 1px solid #444 !important; padding: 3px !important; color: #000000 !important; background-color: #ffffff !important; overflow: hidden !important; text-overflow: ellipsis !important; }
    table.print-table td { white-space: nowrap !important; }
    table.print-table th { background-color: #d5d5d5 !important; font-weight: bold !important; text-align: center !important; white-space: normal !important; word-break: break-word !important; vertical-align: middle !important; -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }
    table.print-table tr { break-inside: avoid !important; page-break-inside: avoid !important; }

    /* LOJAS */
    table.print-loja { font-size: 10px !important; }
    table.print-loja th:nth-child(1), table.print-loja td:nth-child(1) { width: 15% !important; text-align: left !important; }
    table.print-loja th:nth-child(2), table.print-loja td:nth-child(2) { width: 10% !important; text-align: center !important; }
    table.print-loja th:nth-child(3), table.print-loja td:nth-child(3) { width: 50% !important; text-align: left !important; }
    table.print-loja th:nth-child(4), table.print-loja td:nth-child(4) { width: 10% !important; text-align: center !important; }
    table.print-loja th:nth-child(5), table.print-loja td:nth-child(5) { width: 15% !important; text-align: center !important; font-weight: bold !important; background-color: #eeeeee !important; -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }

    /* FORNECEDOR / MATRICIAL */
    table.print-forn { font-size: 9.5px !important; } 
    table.print-forn th:nth-child(1), table.print-forn td:nth-child(1) { width: 8% !important; text-align: center !important; }
    table.print-forn th:nth-child(2), table.print-forn td:nth-child(2) { width: 36% !important; text-align: left !important; }
    table.print-forn th:nth-child(n+3):nth-child(-n+10),
    table.print-forn td:nth-child(n+3):nth-child(-n+10) { width: 7% !important; text-align: center !important; }

    /* SEPARAÇÃO */
    table.print-sep { font-size: 8.5px !important; }
    table.print-sep th:nth-child(1), table.print-sep td:nth-child(1) { width: 14% !important; text-align: left !important; }
    table.print-sep th:nth-child(2), table.print-sep td:nth-child(2) { width: 6%  !important; text-align: center !important; }
    table.print-sep th:nth-child(3), table.print-sep td:nth-child(3) { width: 24% !important; text-align: left !important; }
    table.print-sep th:nth-child(n+4):nth-child(-n+11),
    table.print-sep td:nth-child(n+4):nth-child(-n+11) { width: 7% !important; text-align: center !important; }
}

@media screen {
    #print-section { display: none !important; }
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# FUNÇÃO PARA EXTRAIR APENAS NÚMEROS
# ─────────────────────────────────────────────
def extrair_numero_quantidade(valor):
    if pd.isna(valor) or valor == "":
        return 0
    v_str = str(valor)
    match = re.search(r'\d+', v_str)
    if match:
        return int(match.group())
    return 0

# ─────────────────────────────────────────────
# CONSTANTES E PRODUTOS INICIAIS DA MATRIZ
# ─────────────────────────────────────────────
LOJAS = ["Loja 01", "Loja 02", "Loja 03", "Loja 04", "Loja 05", "Loja 06", "Loja 07", "Loja 08"]
MAPA_LOJAS = {l: l for l in LOJAS}

produtos_iniciais = [
    {"Fornecedor": "RZ Foods ( ITAIQUARA )", "Código": 596686, "Descrição Oficial": "Biscoito Kg Amor Perfeito", "Nome Personalizado": ""},
    {"Fornecedor": "RZ Foods ( ITAIQUARA )", "Código": 596659, "Descrição Oficial": "Biscoito Kg Beijo De Freira", "Nome Personalizado": ""},
    {"Fornecedor": "Coferpan ( 28 Dias )", "Código": 48187, "Descrição Oficial": "Cons Chantilly 907g Richs American Top C/12 und", "Nome Personalizado": ""},
    {"Fornecedor": "Coferpan ( 28 Dias )", "Código": 175553, "Descrição Oficial": "Cons Chantilly 907g Richs Pastry Pride Chocolate C/12", "Nome Personalizado": ""},
    {"Fornecedor": "Harald", "Código": 657190, "Descrição Oficial": "Cons Cobert Gotas 2,050kg Genuine Ao Leite", "Nome Personalizado": ""},
    {"Fornecedor": "Harald", "Código": 657206, "Descrição Oficial": "Cons Cobert Gotas 2,050kg Genuine Meio Amargo", "Nome Personalizado": ""},
    {"Fornecedor": "Reforpan Salware", "Código": 609227, "Descrição Oficial": "Cons Choc Po 33% Salware 1 Kg", "Nome Personalizado": ""},
    {"Fornecedor": "Reforpan Salware", "Código": 609306, "Descrição Oficial": "Cons Gotas Panetone 2,500kg Salware Ao Leite", "Nome Personalizado": ""},
    {"Fornecedor": "Nestle", "Código": 148124, "Descrição Oficial": "Cons Mousse 500g Nestle Suflair Chocolate Sache", "Nome Personalizado": ""},
    {"Fornecedor": "Nestle", "Código": 277259, "Descrição Oficial": "Cons Recheio Cob Moca 2,540kg Chocolate", "Nome Personalizado": ""},
    {"Fornecedor": "TRIPTEM", "Código": 557191, "Descrição Oficial": "Cons Condimento Aglomax Toscana Gkp 30kg", "Nome Personalizado": ""},
    {"Fornecedor": "TRIPTEM", "Código": 560432, "Descrição Oficial": "Cons Emulsificante 505 Duas Rodas 600g", "Nome Personalizado": ""}
]

# ─────────────────────────────────────────────
# FUNÇÃO DE ESTILIZAÇÃO DE EXCEL COM CORES E OBS
# ─────────────────────────────────────────────
def gerar_excel_estilizado(df, sheet_name="Resumo", df_obs=None):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]

        header_fill = PatternFill(start_color='D35400', end_color='D35400', fill_type='solid') 
        alt_row_fill = PatternFill(start_color='FDF2E9', end_color='FDF2E9', fill_type='solid') 
        header_font = Font(color='FFFFFF', bold=True)
        border_style = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column)):
            for cell in row:
                cell.border = border_style
                if row_idx == 0:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    if row_idx % 2 == 0:
                        cell.fill = alt_row_fill
                    
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

        for col in worksheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            worksheet.column_dimensions[col_letter].width = max_length + 2

        if df_obs is not None and not df_obs.empty:
            start_row = worksheet.max_row + 3
            
            cell_title = worksheet.cell(row=start_row, column=1, value="OBSERVAÇÕES GERAIS DAS LOJAS")
            cell_title.font = Font(bold=True, color="D35400")
            
            start_row += 1
            c1 = worksheet.cell(row=start_row, column=1, value="Loja")
            c2 = worksheet.cell(row=start_row, column=2, value="Observação")
            c1.fill = header_fill; c1.font = header_font; c1.border = border_style
            c2.fill = header_fill; c2.font = header_font; c2.border = border_style
            c1.alignment = Alignment(horizontal='center'); c2.alignment = Alignment(horizontal='center')
            
            for idx, row in enumerate(df_obs.itertuples()):
                r = start_row + 1 + idx
                cloja = worksheet.cell(row=r, column=1, value=row.Loja)
                cobs = worksheet.cell(row=r, column=2, value=row.Observacao)
                cloja.border = border_style; cobs.border = border_style
                cloja.alignment = Alignment(horizontal='center', vertical='center')

        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = False
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_margins.left = 0.25
        worksheet.page_margins.right = 0.25
        worksheet.page_margins.top = 0.75
        worksheet.page_margins.bottom = 0.75

    return buffer.getvalue()


# ─────────────────────────────────────────────
# CONEXÃO GOOGLE SHEETS & FUNÇÕES DE DADOS
# ─────────────────────────────────────────────
conn = st.connection("gsheets", type=GSheetsConnection)

WS_PRODUTOS = "PadConf_Produtos"
WS_PEDIDOS  = "PadConf_Pecas"
WS_OBS      = "PadConf_Obs"

def parse_bool(x):
    if isinstance(x, bool): return x
    if isinstance(x, (int, float)): return bool(x) and not pd.isna(x)
    return str(x).strip().upper() in ['TRUE', 'VERDADEIRO', '1', 'V', 'SIM', 'YES', 'T', 'X']

@st.cache_data(ttl=15)
def carregar_obs():
    try:
        df_obs = conn.read(worksheet=WS_OBS, ttl=0)
    except Exception:
        df_obs = pd.DataFrame(columns=["Loja", "Observacao"])

    if df_obs.empty or "Loja" not in df_obs.columns:
        df_obs = pd.DataFrame({"Loja": LOJAS, "Observacao": [""] * len(LOJAS)})
        try:
            conn.update(worksheet=WS_OBS, data=df_obs)
        except Exception:
            pass 
            
    if "Observacao" not in df_obs.columns:
        df_obs["Observacao"] = ""
        
    df_obs["Observacao"] = df_obs["Observacao"].fillna("").astype(str)

    for loja in LOJAS:
        if loja not in df_obs["Loja"].values:
            nova_obs = pd.DataFrame([{"Loja": loja, "Observacao": ""}])
            df_obs = pd.concat([df_obs, nova_obs], ignore_index=True)
            
    return df_obs

def salvar_obs(df_obs_to_save):
    conn.update(worksheet=WS_OBS, data=df_obs_to_save)
    st.cache_data.clear()

@st.cache_data(ttl=15)
def carregar_catalogo_padconf():
    try:
        df = conn.read(worksheet=WS_PRODUTOS, ttl=0, usecols=list(range(20)))
    except ValueError as e:
        if "Spreadsheet must be specified" in str(e):
            st.error("🚨 **Erro Crítico:** URL da Planilha não especificada nas configurações do Streamlit Cloud (Secrets).")
            st.stop()
        else:
            raise e
    except Exception as e:
        st.error(f"Erro ao conectar na aba {WS_PRODUTOS}: {e}")
        st.stop()

    if df.empty or "Fornecedor" not in df.columns:
        df_init = pd.DataFrame(produtos_iniciais)
        for loja in LOJAS: df_init[loja] = True
        conn.update(worksheet=WS_PRODUTOS, data=df_init)
        df = df_init.copy()

    need_update = False

    if "Descrição" in df.columns and "Descrição Oficial" not in df.columns:
        df = df.rename(columns={"Descrição": "Descrição Oficial"})
        need_update = True

    if "Nome Personalizado" not in df.columns:
        df["Nome Personalizado"] = ""
        need_update = True

    if need_update:
        conn.update(worksheet=WS_PRODUTOS, data=df.drop(columns=["Descrição"], errors="ignore"))

    novas_colunas = {}
    for col in df.columns:
        col_str = str(col).strip()
        for loja in LOJAS:
            if loja.lower() in col_str.lower():
                novas_colunas[col] = loja
    df = df.rename(columns=novas_colunas)

    for loja in LOJAS:
        if loja not in df.columns:
            df[loja] = False
        else:
            df[loja] = df[loja].apply(parse_bool)

    if "Código" in df.columns:
        df["Código"] = pd.to_numeric(df["Código"], errors='coerce').fillna(0).astype(int)

    def obter_nome_final(row):
        apelido = str(row.get("Nome Personalizado", "")).strip()
        if apelido and apelido.lower() != "nan":
            return apelido
        return str(row.get("Descrição Oficial", "")).strip()

    df["Descrição"] = df.apply(obter_nome_final, axis=1)

    return df

@st.cache_data(ttl=15)
def carregar_pedidos():
    try:
        df_pedidos = conn.read(worksheet=WS_PEDIDOS, ttl=0)
    except ValueError as e:
        if "Spreadsheet must be specified" in str(e):
            st.error("🚨 **Erro Crítico:** URL da Planilha não especificada nas configurações.")
            st.stop()
        else:
            raise e
    except Exception as e:
        st.error(f"Erro ao conectar na aba {WS_PEDIDOS}: {e}")
        st.stop()

    df_cat = carregar_catalogo_padconf()

    if df_pedidos.empty or "Código" not in df_pedidos.columns:
        df_init = df_cat[["Fornecedor", "Código", "Descrição"]].copy()
        for loja in LOJAS:
            df_init[loja] = "" 
        if not df_init.empty:
            conn.update(worksheet=WS_PEDIDOS, data=df_init)
        return df_init

    df_pedidos = df_pedidos.drop(columns=["Descrição", "Fornecedor"], errors="ignore")
    df_pedidos = pd.merge(df_cat[["Código", "Fornecedor", "Descrição"]], df_pedidos, on="Código", how="left")

    if "Código" in df_pedidos.columns:
        df_pedidos["Código"] = pd.to_numeric(df_pedidos["Código"], errors='coerce').fillna(0).astype(int)

    for loja in LOJAS:
        if loja in df_pedidos.columns:
            df_pedidos[loja] = df_pedidos[loja].fillna("").astype(str).apply(lambda x: "" if str(x).strip() in ["0", "0.0", "0,0"] else x)
        else:
            df_pedidos[loja] = ""

    return df_pedidos

def salvar_pedidos(df_to_save):
    conn.update(worksheet=WS_PEDIDOS, data=df_to_save)
    st.cache_data.clear()

def salvar_catalogo(df_to_save):
    df_clean = df_to_save.drop(columns=["Descrição"], errors="ignore")
    conn.update(worksheet=WS_PRODUTOS, data=df_clean)
    st.cache_data.clear()

# ─────────────────────────────────────────────
# SISTEMA DE LOGIN
# ─────────────────────────────────────────────
if st.session_state['usuario_logado_padconf'] is None:
    st.write("<br><br>", unsafe_allow_html=True)
    _, col2, _ = st.columns([1, 1.4, 1])
    with col2:
        with st.container(border=True):
            h1, h2 = st.columns([4, 1])
            with h1:
                st.markdown("""
                    <h2 style='margin-bottom:0'>Portal de Pedidos</h2>
                    <p style='color:#7d8590;font-size:14px;margin-top:4px'>Padaria e Confeitaria — Molicenter</p>
                """, unsafe_allow_html=True)
            with h2:
                st.write("")
                try:
                    st.image("passaro_logo.png", width=60)
                except Exception:
                    st.markdown("🥐", unsafe_allow_html=True)

            st.divider()
            usuarios_permitidos = ["Selecione..."] + ["Administrador"] + LOJAS
            usuario_selecionado = st.selectbox("👤 Usuário de acesso:", usuarios_permitidos)

            senha_digitada = st.text_input("🔑 Senha de acesso:", type="password", autocomplete="off")

            st.write("<br>", unsafe_allow_html=True)

            if st.button("Entrar no Sistema", type="primary", use_container_width=True):
                if usuario_selecionado == "Selecione...":
                    st.error("⚠️ Por favor, selecione um usuário.")
                elif usuario_selecionado == "Administrador" and senha_digitada == "moli0000":
                    st.session_state['usuario_logado_padconf'] = usuario_selecionado
                    st.rerun()
                elif usuario_selecionado in LOJAS and senha_digitada == "moli1234":
                    st.session_state['usuario_logado_padconf'] = usuario_selecionado
                    st.rerun()
                elif senha_digitada:
                    st.error("⚠️ Senha incorreta. Tente novamente.")

            st.markdown('<p style="font-size: 11px; color: #7d8590; text-align: center; margin-top: 10px;">🔒 Acesso restrito — Molicenter © 2026</p>', unsafe_allow_html=True)
    st.stop()

# ─────────────────────────────────────────────
# PÓS-LOGIN
# ─────────────────────────────────────────────
usuario_atual = st.session_state['usuario_logado_padconf']
acesso_total  = usuario_atual == "Administrador"

if not acesso_total:
    st.markdown("""
    <style>
        section[data-testid="stSidebar"] { display: none !important; }
        [data-testid="collapsedControl"]  { display: none !important; }
        .main .block-container { max-width: 100% !important; padding-left: 2.5rem !important; padding-right: 2.5rem !important; }
    </style>
    """, unsafe_allow_html=True)

# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    try:
        st.image("passaro_logo.png", width=72)
    except Exception:
        st.markdown("🥐")

    st.markdown(f"### Olá, *{usuario_atual}*")
    st.caption("Sistema de Pedidos — Padaria e Confeitaria")
    st.divider()

    if acesso_total:
        perfil_navegacao = st.radio("📍 Navegação:", [
            "Visão por Fornecedor (Resumo)",
            "Separação e Fechamento",
            "Visão das Lojas",
            "Catálogo de Produtos"
        ])
    else:
        perfil_navegacao = "Visão das Lojas"

    st.divider()

    df_ped = carregar_pedidos()
    if not df_ped.empty and set(LOJAS).issubset(df_ped.columns):
        temp_sum = df_ped[LOJAS].map(extrair_numero_quantidade)
        total_preenchidos = (temp_sum > 0).any(axis=1).sum()
    else:
        total_preenchidos = 0

    st.metric("Itens c/ pedido", total_preenchidos, help="Itens com ao menos 1 quantidade preenchida")

    st.divider()

    if st.button("🔄 Sincronizar Dados", use_container_width=True):
        st.cache_data.clear()
        st.session_state['reset_counter_padconf'] += 1
        st.rerun()

    st.write("<br>", unsafe_allow_html=True)

    if st.button("🚪 Sair / Logout", use_container_width=True):
        st.session_state['usuario_logado_padconf'] = None
        st.rerun()

# ─────────────────────────────────────────────
# FUNÇÃO MODAL DE CONFIRMAÇÃO PARA ZERAR
# ─────────────────────────────────────────────
@st.dialog("🚨 Confirmação Necessária")
def modal_zerar_pedidos():
    st.markdown("Tem certeza que deseja **zerar todos os pedidos e observações** de todas as lojas?")
    st.markdown("⚠️ *Esta ação irá limpar as quantidades diretamente no Google Sheets.*")

    st.write("<br>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        if st.button("❌ Não, cancelar", use_container_width=True):
            st.rerun()
    with c2:
        if st.button("✔️ Sim, zerar tudo", type="primary", use_container_width=True):
            st.session_state['reset_counter_padconf'] += 1
            df_main = carregar_pedidos()
            for loja in LOJAS:
                if loja in df_main.columns:
                    df_main[loja] = ""
            salvar_pedidos(df_main)
            
            df_obs = carregar_obs()
            df_obs["Observacao"] = ""
            salvar_obs(df_obs)
            
            st.rerun()

# ─────────────────────────────────────────────
# ROTA 1 — SEPARAÇÃO E FECHAMENTO (Admin)
# ─────────────────────────────────────────────
if perfil_navegacao == "Separação e Fechamento":
    st.markdown("""
    <div class="page-header hide-print" style="background: linear-gradient(90deg, var(--brown-dark) 0%, #1c140d 100%); padding: 14px 20px; border-radius: 10px; margin-bottom: 22px;">
        <span style="font-size: 26px; margin-right: 12px;">📊</span>
        <div style="display: inline-block; vertical-align: top;">
            <div style="font-size: 20px; font-weight: 700; color: var(--text-header);">Separação e Fechamento — Padaria e Confeitaria</div>
            <div style="font-size: 12px; color: var(--text-muted); margin-top: 2px;">Consolidado geral de quantidades por Categoria e Código</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    with st.container(border=True):
        df_base = carregar_pedidos()

        if df_base.empty:
            st.warning("A base de pedidos está vazia. Cadastre produtos no Catálogo primeiro.")
            st.stop()

        cols_order = ["Fornecedor", "Código", "Descrição"] + LOJAS
        df_exibir = df_base[cols_order]

        termo_busca_sep = st.text_input("🔍 Buscar Produto (por Código ou Nome):", placeholder="Digite aqui para filtrar a lista abaixo...", key="busca_sep")
        if termo_busca_sep:
            mask = df_exibir["Descrição"].str.contains(termo_busca_sep, case=False, na=False) | \
                   df_exibir["Código"].astype(str).str.contains(termo_busca_sep, case=False, na=False)
            df_exibir = df_exibir[mask]
            st.caption("⚠️ *Aviso: Salve suas alterações antes de limpar a busca, para não perder as edições atuais.*")

        altura_dinamica = min(600, max(100, int(len(df_exibir) * 35.5) + 42))

        col_cfg = {
            "Fornecedor":  st.column_config.TextColumn("Categoria", disabled=True),
            "Código":      st.column_config.NumberColumn("Cód.", width=80, format="%d", disabled=True),
            "Descrição":   st.column_config.TextColumn("Produto", disabled=True),
        }
        for loja, novo_nome in MAPA_LOJAS.items():
            col_cfg[loja] = st.column_config.TextColumn(novo_nome)

        # -------------------------------------------------------------
        # OTIMIZAÇÃO: Formulário para segurar as edições na memória
        # -------------------------------------------------------------
        with st.form(key=f"form_sep_{st.session_state['reset_counter_padconf']}"):
            df_editado = st.data_editor(
                df_exibir,
                hide_index=True,
                use_container_width=True,
                height=altura_dinamica,
                column_config=col_cfg,
                key=f"sep_editor_{st.session_state['reset_counter_padconf']}"
            )
            
            st.write("<br>", unsafe_allow_html=True)
            col_espaco, col_salvar = st.columns([7, 3])
            with col_salvar:
                btn_salvar_sep = st.form_submit_button("💾 Salvar Alterações", type="primary", use_container_width=True)
        
        if btn_salvar_sep:
            df_to_save = carregar_pedidos()
            for _, row_edit in df_editado.iterrows():
                mask = (df_to_save["Fornecedor"] == row_edit["Fornecedor"]) & (df_to_save["Código"] == row_edit["Código"])
                for loja in LOJAS:
                    df_to_save.loc[mask, loja] = str(row_edit[loja])
            
            salvar_pedidos(df_to_save)
            st.success("✅ Pedidos salvos na nuvem com sucesso!")
            st.rerun()

        html_table = df_editado.to_html(index=False, classes=["print-table", "print-sep"])
        st.markdown(f"""<div id="print-section">
<h2 style="color: black; margin-bottom: 10px; text-align: center; border-bottom: 2px solid black; padding-bottom: 5px;">
    Resumo de Separação — Padaria e Confeitaria
</h2>
<div class="print-container">
{html_table}
</div>
</div>""", unsafe_allow_html=True)

        st.divider()
        col_csv, col_excel, col_print, col_zerar = st.columns([1.5, 1.5, 1.5, 2.5])

        df_obs_print = carregar_obs()
        df_obs_filtrado = df_obs_print[df_obs_print["Observacao"].str.strip() != ""]

        with col_csv:
            df_csv = df_editado.copy().rename(columns=MAPA_LOJAS)
            csv = df_csv.to_csv(index=False).encode("utf-8")
            st.download_button("⬇️ Exportar CSV", data=csv, file_name="separacao_padaria_confeitaria.csv", mime="text/csv", use_container_width=True)

        with col_excel:
            df_exp = df_editado.copy().rename(columns=MAPA_LOJAS)
            excel_data = gerar_excel_estilizado(df_exp, "Separação Padaria", df_obs=df_obs_filtrado)
            st.download_button("⬇️ Exportar Excel", data=excel_data, file_name="separacao_padaria_confeitaria.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

        with col_print:
            if st.button("🖨️ Imprimir Visualização", use_container_width=True):
                components.html(
                    "<script>"
                    "var s=document.createElement('style');"
                    "s.id='__sep_land__';"
                    "s.innerHTML='@media print { @page { size: landscape; margin: 5mm; } html,body,.stApp,#root,[data-testid=\"stAppViewContainer\"],[data-testid=\"stMainBlockContainer\"],.main,.block-container { padding: 0 !important; margin: 0 !important; } }';"
                    "window.parent.document.head.appendChild(s);"
                    "window.parent.print();"
                    "setTimeout(function(){"
                    "var e=window.parent.document.getElementById('__sep_land__');"
                    "if(e)e.remove();"
                    "},3000);"
                    "</script>",
                    height=0
                )

        with col_zerar:
            if st.button("🚨 Zerar Todos os Pedidos", use_container_width=True, key="btn_zerar_sep_padconf"):
                modal_zerar_pedidos()
                
        st.write("<br>", unsafe_allow_html=True)
        st.markdown("### 📝 Observações das Lojas")
        if not df_obs_filtrado.empty:
            st.dataframe(df_obs_filtrado, hide_index=True, use_container_width=True)
        else:
            st.info("Nenhuma observação registrada nesta semana.")

# ─────────────────────────────────────────────
# ROTA 2 — VISÃO DAS LOJAS
# ─────────────────────────────────────────────
elif perfil_navegacao == "Visão das Lojas":
    if acesso_total:
        loja_selecionada = st.selectbox("👁️ Visualizar como:", LOJAS)
    else:
        loja_selecionada = usuario_atual

    col_info, col_logout = st.columns([8, 2])
    with col_info:
        id_loja = MAPA_LOJAS.get(loja_selecionada, loja_selecionada)
        st.markdown(f"""
        <div class="topbar-loja hide-print">
            <div class="topbar-left">
                <span style="font-size:22px">🥐</span>
                <div>
                    <div class="topbar-title">{loja_selecionada} — Padaria e Confeitaria <span class="erp-badge">🟢 Conectado ao ERP</span></div>
                    <div class="topbar-sub">Preencha a quantidade dos produtos e use o campo final para observações gerais.</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    with col_logout:
        st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)
        if st.button("🚪 Sair / Logout", use_container_width=True):
            st.session_state['usuario_logado_padconf'] = None
            st.rerun()

    df_cat = carregar_catalogo_padconf()
    df_cat_loja = df_cat[df_cat[loja_selecionada] == True].copy()

    if df_cat_loja.empty:
        st.warning(f"Nenhum produto habilitado para a {loja_selecionada} no momento.")
        st.stop()

    df_all = carregar_pedidos()
    
    df_loja_view = pd.merge(
        df_cat_loja[["Fornecedor", "Código", "Descrição"]],
        df_all[["Fornecedor", "Código", loja_selecionada]],
        on=["Fornecedor", "Código"],
        how="left"
    )
    df_loja_view[loja_selecionada] = df_loja_view[loja_selecionada].fillna("").astype(str)
    
    df_loja_view = df_loja_view.rename(columns={loja_selecionada: "Qtde"})

    try:
        conn_pg = st.connection("banco_erp", type="sql")
        mapa_banco_erp = {
            "Loja 01": "001", "Loja 02": "002", "Loja 03": "003",
            "Loja 04": "004", "Loja 05": "005", "Loja 06": "006",
            "Loja 07": "007", "Loja 08": "008"
        }
        cod_empresa_banco = mapa_banco_erp.get(loja_selecionada, "001")

        query_erp = f"""
            SELECT cade_codempresa,
                   cade_codigo,
                   cadp_descricao,
                   estoque
            FROM "python_estoque"
            WHERE cade_codempresa::text = '{cod_empresa_banco}'
            ORDER BY cade_codempresa, cade_codigo
        """
        df_erp = conn_pg.query(query_erp, ttl=300)

        if not df_erp.empty:
            df_erp = df_erp.rename(columns={"cade_codigo": "Código", "estoque": "Estoque"})
            df_loja_view = pd.merge(df_loja_view, df_erp[["Código", "Estoque"]], on="Código", how="left")
        else:
            df_loja_view["Estoque"] = 0

    except Exception as e:
        if "No database configured" in str(e) or "missing" in str(e).lower():
             st.error("⚠️ Aviso: As credenciais do banco_erp também precisam estar nos Secrets do Streamlit para puxar o estoque.")
        else:
             st.error(f"⚠️ Erro ao puxar dados do ERP PostgreSQL: {e}")
        df_loja_view["Estoque"] = 0

    df_loja_view["Estoque"] = df_loja_view["Estoque"].fillna(0).astype(int)
    df_loja_view = df_loja_view[["Fornecedor", "Código", "Descrição", "Estoque", "Qtde"]]

    with st.container(border=True):
        termo_busca_loja = st.text_input("🔍 Buscar Produto (por Código ou Nome):", placeholder="Digite aqui para filtrar a lista abaixo...", key="busca_lojas")
        if termo_busca_loja:
            mask = df_loja_view["Descrição"].str.contains(termo_busca_loja, case=False, na=False) | \
                   df_loja_view["Código"].astype(str).str.contains(termo_busca_loja, case=False, na=False)
            df_loja_view = df_loja_view[mask]
            st.caption("⚠️ *Aviso: Salve o seu pedido antes de limpar ou alterar a busca, para não perder o que foi digitado.*")

        altura_dinamica = min(600, max(100, int(len(df_loja_view) * 35.5) + 42))

        col_cfg_loja = {
            "Fornecedor": st.column_config.TextColumn("Categoria", width=130, disabled=True),
            "Código":     st.column_config.NumberColumn("Cód.", width=80, format="%d", disabled=True),
            "Descrição":  st.column_config.TextColumn("Produto", width=300, disabled=True),
            "Estoque":    st.column_config.NumberColumn("📦 Estoque", width=100, format="%d", disabled=True),
            "Qtde":       st.column_config.TextColumn("🛒 Qtde", width=120),
        }

        df_obs_atual = carregar_obs()
        obs_existente = df_obs_atual.loc[df_obs_atual["Loja"] == loja_selecionada, "Observacao"].values
        obs_texto_inicial = obs_existente[0] if len(obs_existente) > 0 else ""

        # -------------------------------------------------------------
        # OTIMIZAÇÃO: Formulário da Loja para prevenir travamento a cada dígito
        # -------------------------------------------------------------
        with st.form(key=f"form_loja_{loja_selecionada}_{st.session_state['reset_counter_padconf']}"):
            df_editado = st.data_editor(
                df_loja_view,
                column_config=col_cfg_loja,
                hide_index=True,
                use_container_width=True,
                height=altura_dinamica,
                key=f"loja_editor_{st.session_state['reset_counter_padconf']}"
            )
            
            st.write("<br>", unsafe_allow_html=True)
            nova_obs = st.text_area("📝 Observação Geral da Loja", value=obs_texto_inicial, height=80, placeholder="Caso necessite, digite alguma observação geral para esse pedido...")

            # Cálculos em tempo real baseados nos dados *já salvos ou enviados na última interação*
            df_editado_nums = df_editado["Qtde"].apply(extrair_numero_quantidade)
            itens_com_pedido = int((df_editado_nums > 0).sum())
            total_itens      = len(df_loja_view) 
            total_unidades   = int(df_editado_nums.sum())
            pct              = round(itens_com_pedido / total_itens * 100) if total_itens else 0

            st.divider()
            m1, m2, m3, col_btn = st.columns([2.5, 2.2, 1.8, 3.5])
            
            texto_itens = f"{itens_com_pedido} / {total_itens}" if not termo_busca_loja else f"{itens_com_pedido} / {total_itens} (na busca)"
            with m1: st.metric("Itens preenchidos", texto_itens)
            with m2: st.metric("Total de unidades", total_unidades)
            with m3: st.metric("Cobertura", f"{pct}%")

            with col_btn:
                st.write("<br>", unsafe_allow_html=True)
                btn_salvar_loja = st.form_submit_button("💾 Salvar Pedido e OBS", type="primary", use_container_width=True)

        if btn_salvar_loja:
            df_main = carregar_pedidos()
            for _, row in df_editado.iterrows():
                mask = (df_main["Fornecedor"] == row["Fornecedor"]) & (df_main["Código"] == row["Código"])
                if mask.any():
                    df_main.loc[mask, loja_selecionada] = str(row["Qtde"])
                else:
                    nova_linha = {"Fornecedor": row["Fornecedor"], "Código": row["Código"], "Descrição": row["Descrição"]}
                    for l in LOJAS: 
                        nova_linha[l] = ""
                    nova_linha[loja_selecionada] = str(row["Qtde"])
                    df_main = pd.concat([df_main, pd.DataFrame([nova_linha])], ignore_index=True)
            salvar_pedidos(df_main)
            
            df_obs_save = carregar_obs()
            mask_obs = df_obs_save["Loja"] == loja_selecionada
            if mask_obs.any():
                df_obs_save.loc[mask_obs, "Observacao"] = nova_obs
            else:
                df_obs_save = pd.concat([df_obs_save, pd.DataFrame([{"Loja": loja_selecionada, "Observacao": nova_obs}])], ignore_index=True)
            salvar_obs(df_obs_save)
            
            st.success(f"✅ Pedido da {loja_selecionada} enviado para a nuvem com sucesso!")
            st.rerun()

        # ÁREA DE IMPRESSÃO - Fica de fora do form
        df_imprimir = df_editado.copy()
        df_imprimir = df_imprimir.rename(columns={"Estoque": "Est.", "Qtde": "Ped."})
        html_table_loja = df_imprimir.to_html(index=False, classes=["print-table", "print-loja"])

        st.markdown(f"""<div id="print-section">
<h2 style="color: black; margin-bottom: 10px; text-align: center; border-bottom: 2px solid black; padding-bottom: 5px;">
    Resumo do Pedido — {loja_selecionada}
</h2>
<div class="print-container">
{html_table_loja}
</div>
<br>
<b>OBSERVAÇÃO GERAL:</b> {nova_obs}
</div>""", unsafe_allow_html=True)

        col_print_loja, _ = st.columns([2, 8])
        with col_print_loja:
            if st.button("🖨️ Imprimir Visualização", use_container_width=True):
                components.html(
                    "<script>"
                    "var s=document.createElement('style');"
                    "s.id='__loja_print__';"
                    "s.innerHTML='@media print { @page { size: landscape; margin: 5mm; } html,body,.stApp,#root,[data-testid=\"stAppViewContainer\"],[data-testid=\"stMainBlockContainer\"],.main,.block-container { padding: 0 !important; margin: 0 !important; } }';"
                    "window.parent.document.head.appendChild(s);"
                    "window.parent.print();"
                    "setTimeout(function(){var e=window.parent.document.getElementById('__loja_print__');if(e)e.remove();},3000);"
                    "</script>", 
                    height=0
                )

# ─────────────────────────────────────────────
# ROTA 3 — VISÃO POR FORNECEDOR / RESUMO (Admin)
# ─────────────────────────────────────────────
elif perfil_navegacao == "Visão por Fornecedor (Resumo)":
    st.markdown("""
    <div class="hide-print" style="background: linear-gradient(90deg, var(--brown-dark) 0%, #1c140d 100%); padding: 14px 20px; border-radius: 10px; margin-bottom: 22px;">
        <span style="font-size: 26px; margin-right: 12px;">🥐</span>
        <div style="display: inline-block; vertical-align: top;">
            <div style="font-size: 20px; font-weight: 700; color: var(--text-header);">Visão por Categoria — Padaria e Confeitaria</div>
            <div style="font-size: 12px; color: var(--text-muted); margin-top: 2px;">Resumo consolidado agrupado pelas categorias de produtos</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    df_all = carregar_pedidos()
    df_cat = carregar_catalogo_padconf()

    if df_all.empty or df_cat.empty:
        st.warning("Não há dados de pedidos ou catálogo preenchidos.")
        st.stop()

    nomes_forn = df_cat["Fornecedor"].dropna().unique().tolist()
    html_print_content = ""

    for i in range(0, len(nomes_forn), 1):
        cols = st.columns(1, gap="small")
        for j, fornecedor in enumerate(nomes_forn[i:i+1]):

            df_forn = df_all[df_all["Fornecedor"] == fornecedor].copy()
            colunas_presentes = LOJAS
            
            df_forn_view = df_forn[["Código", "Descrição"] + colunas_presentes].copy()
            
            lojas_renomeadas = {l: MAPA_LOJAS[l] for l in colunas_presentes}
            df_forn_view = df_forn_view.rename(columns=lojas_renomeadas)
            lojas_cols_renomeadas = [MAPA_LOJAS[l] for l in colunas_presentes]

            col_cfg_forn = {
                "Código":    st.column_config.NumberColumn("Cód.", width=80, format="%d", disabled=True),
                "Descrição": st.column_config.TextColumn("Produto", disabled=False),
            }
            for c in lojas_cols_renomeadas:
                col_cfg_forn[c] = st.column_config.TextColumn(c, disabled=False)

            altura = min(600, max(100, int(len(df_forn_view) * 35.5) + 42))

            with cols[j]:
                with st.container(border=True):
                    # -------------------------------------------------------------
                    # OTIMIZAÇÃO: Isolando as visões em formulários
                    # -------------------------------------------------------------
                    with st.form(key=f"form_forn_{fornecedor}_{st.session_state['reset_counter_padconf']}"):
                        st.markdown('<div class="title-input">', unsafe_allow_html=True)
                        st.text_input(
                            "Categoria",
                            value=f"🥐 {fornecedor}",
                            label_visibility="collapsed",
                            key=f"title_forn_{fornecedor}_{st.session_state['reset_counter_padconf']}"
                        )
                        st.markdown('</div>', unsafe_allow_html=True)

                        cols_order_forn = ["Código", "Descrição"] + lojas_cols_renomeadas
                        df_forn_edit = st.data_editor(
                            df_forn_view[cols_order_forn],
                            hide_index=True,
                            use_container_width=True,
                            column_config=col_cfg_forn,
                            height=altura,
                            num_rows="fixed",
                            key=f"forn_editor_{fornecedor}_{st.session_state['reset_counter_padconf']}"
                        )

                        df_forn_nums = df_forn_view[lojas_cols_renomeadas].map(extrair_numero_quantidade)
                        total_geral = int(df_forn_nums.sum().sum())
                        
                        st.markdown(f"""
                            <div style="text-align:right; font-weight:700; margin-top:6px; color:var(--brown-bright); font-size:15px;">
                                Total Geral: {total_geral} unidades
                            </div>
                        """, unsafe_allow_html=True)
                        
                        # Botão Dummy/Atualizar para não quebrar a lógica visual que vocês já tinham
                        col_espaco, col_btn = st.columns([8, 2])
                        with col_btn:
                            st.form_submit_button("Atualizar Visão", use_container_width=True)

                    html_table = df_forn_edit.to_html(index=False, classes=["print-table", "print-forn"])
                    for loja in LOJAS:
                        partes = loja.split(" ")
                        if len(partes) == 2:
                            html_table = html_table.replace(
                                f"<th>{loja}</th>",
                                f"<th>{partes[0]}<br>{partes[1]}</th>"
                            )
                    html_print_content += f"<h3 style='color: black; margin-top: 10px; margin-bottom: 4px;'>🥐 {fornecedor}</h3>\n"
                    html_print_content += f"{html_table}\n"
                    html_print_content += f"<div style='text-align:right; font-weight:bold; font-size:11px; margin-top:3px; margin-bottom: 8px; color: black;'>Total da Categoria: {total_geral} unidades</div>\n"

        st.write("<br>", unsafe_allow_html=True)

    st.markdown(f"""<div id="print-section">
<h2 style="color: black; margin-bottom: 10px; text-align: center; border-bottom: 2px solid black; padding-bottom: 5px;">
    Visão por Categoria (Resumo) — Padaria e Confeitaria
</h2>
<div class="print-container">
{html_print_content}
</div>
</div>""", unsafe_allow_html=True)

    st.divider()
    
    col_csv, col_excel, col_space, col_print, col_zerar = st.columns([1.5, 1.5, 2.0, 2.5, 2.5])
    
    df_obs_print = carregar_obs()
    df_obs_filtrado = df_obs_print[df_obs_print["Observacao"].str.strip() != ""]
    
    df_export = pd.DataFrame()
    for forn in nomes_forn:
        df_f = df_all[df_all["Fornecedor"] == forn].copy()
        if not df_f.empty:
            df_export = pd.concat([df_export, df_f], ignore_index=True)
            
    if not df_export.empty:
        cols_final_export = ["Código", "Descrição", "Fornecedor"] + LOJAS
        df_export = df_export[cols_final_export]
        df_export = df_export.rename(columns=MAPA_LOJAS)

    with col_csv:
        if not df_export.empty:
            csv_str = df_export.to_csv(index=False).encode('utf-8')
            st.download_button("⬇️ Exportar CSV", data=csv_str, file_name="visao_categoria_padconf.csv", mime="text/csv", use_container_width=True)

    with col_excel:
        if not df_export.empty:
            excel_data = gerar_excel_estilizado(df_export, "Visão Categorias", df_obs=df_obs_filtrado)
            st.download_button("⬇️ Exportar Excel", data=excel_data, file_name="visao_categoria_padconf.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    with col_print:
        if st.button("🖨️ Imprimir Resumo Geral", use_container_width=True):
            components.html(
                "<script>"
                "var s=document.createElement('style');"
                "s.id='__forn_port__';"
                "s.innerHTML='@media print { @page { size: landscape; margin: 5mm; } html,body,.stApp,#root,[data-testid=\"stAppViewContainer\"],[data-testid=\"stMainBlockContainer\"],.main,.block-container { padding: 0 !important; margin: 0 !important; } }';"
                "window.parent.document.head.appendChild(s);"
                "window.parent.print();"
                "setTimeout(function(){"
                "var e=window.parent.document.getElementById('__forn_port__');"
                "if(e)e.remove();"
                "},3000);"
                "</script>",
                height=0
            )

    with col_zerar:
        if st.button("🚨 Zerar Todos os Pedidos", use_container_width=True, key="btn_zerar_forn_padconf"):
            modal_zerar_pedidos()
            
    st.write("<br>", unsafe_allow_html=True)
    st.markdown("### 📝 Observações das Lojas")
    if not df_obs_filtrado.empty:
        st.dataframe(df_obs_filtrado, hide_index=True, use_container_width=True)
    else:
        st.info("Nenhuma observação registrada nesta semana.")

# ─────────────────────────────────────────────
# ROTA 4 — CATÁLOGO DE PRODUTOS
# ─────────────────────────────────────────────
elif perfil_navegacao == "Catálogo de Produtos":
    st.markdown("""
    <div class="page-header hide-print" style="background: linear-gradient(90deg, var(--brown-dark) 0%, #1c140d 100%); padding: 14px 20px; border-radius: 10px; margin-bottom: 22px;">
        <span style="font-size: 26px; margin-right: 12px;">🗂️</span>
        <div style="display: inline-block; vertical-align: top;">
            <div style="font-size: 20px; font-weight: 700; color: var(--text-header);">Catálogo de Padaria e Confeitaria</div>
            <div style="font-size: 12px; color: var(--text-muted); margin-top: 2px;">Atualize nomes direto do ERP ou crie apelidos personalizados. Os apelidos terão prioridade em todo o sistema. <br><b>Dica: Para adicionar todos os produtos rapidamente, cole-os direto na sua planilha do Google Sheets.</b></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    df_catalogo = carregar_catalogo_padconf()
    df_editor_input = df_catalogo.drop(columns=["Descrição"], errors="ignore")

    if "Fornecedor" in df_editor_input.columns:
        df_editor_input["Fornecedor"] = df_editor_input["Fornecedor"].astype(str).str.strip()

    for loja in LOJAS:
        if loja in df_editor_input.columns:
            df_editor_input[loja] = df_editor_input[loja].fillna(False).astype(bool)

    df_editor_input["Código"] = pd.to_numeric(df_editor_input["Código"], errors='coerce').fillna(0).astype(int)

    cols_texto = ["Fornecedor", "Descrição Oficial", "Nome Personalizado"]
    for col in cols_texto:
        if col in df_editor_input.columns:
            df_editor_input[col] = df_editor_input[col].fillna("").astype(str)

    ordem_colunas = ["Fornecedor", "Código", "Descrição Oficial", "Nome Personalizado"] + LOJAS
    df_editor_input = df_editor_input[ordem_colunas]
    
    termo_busca_cat = st.text_input("🔍 Buscar Produto:", placeholder="Filtrar por código, descrição ou categoria...", key="busca_cat")
    if termo_busca_cat:
        mask = df_editor_input["Descrição Oficial"].str.contains(termo_busca_cat, case=False, na=False) | \
               df_editor_input["Código"].astype(str).str.contains(termo_busca_cat, case=False, na=False) | \
               df_editor_input["Fornecedor"].str.contains(termo_busca_cat, case=False, na=False)
        df_editor_input = df_editor_input[mask]
        
    altura_dinamica_cat = min(600, max(100, int(len(df_editor_input) * 35.5) + 42))

    with st.container(border=True):
        categorias_banco = df_editor_input["Fornecedor"].unique().tolist()
        categorias_padrao = [
            "RZ Foods ( ITAIQUARA )", "Coferpan ( 28 Dias )", "Harald", 
            "Reforpan Salware", "Nestle", "TRIPTEM", "675576 - Peniel", "CD"
        ]
        
        opcoes_forn = sorted(list(set([c for c in (categorias_banco + categorias_padrao) if c.strip() != ""])))

        col_cfg_cat = {
            "Fornecedor":         st.column_config.SelectboxColumn("Categoria", options=opcoes_forn, width=200, required=True),
            "Código":             st.column_config.NumberColumn("Cód.", width=80, format="%d", required=True),
            "Descrição Oficial":  st.column_config.TextColumn("Nome Oficial (ERP)", width=280, disabled=False),
            "Nome Personalizado": st.column_config.TextColumn("Nome Personalizado (Apelido)", width=230),
        }
        for loja in LOJAS:
            col_cfg_cat[loja] = st.column_config.CheckboxColumn(loja, default=False)

        # -------------------------------------------------------------
        # OTIMIZAÇÃO: Isolando a edição do catálogo
        # -------------------------------------------------------------
        with st.form(key=f"form_cat_{st.session_state['reset_counter_padconf']}"):
            edited_cat = st.data_editor(
                df_editor_input,
                use_container_width=True,
                hide_index=True,
                height=altura_dinamica_cat,
                num_rows="dynamic",
                column_config=col_cfg_cat,
                key=f"editor_catalogo_padconf_{st.session_state['reset_counter_padconf']}"
            )
            
            st.write("<br>", unsafe_allow_html=True)
            col_espaco, col_atualizar = st.columns([7, 3])
            with col_atualizar:
                btn_salvar_cat = st.form_submit_button("💾 Salvar Catálogo", type="primary", use_container_width=True)

        if btn_salvar_cat:
            df_to_save = carregar_catalogo_padconf().drop(columns=["Descrição"], errors="ignore")
            
            for _, row in edited_cat.iterrows():
                mask = (df_to_save["Código"] == row["Código"]) & (df_to_save["Fornecedor"] == row["Fornecedor"])
                if mask.any():
                    for c in df_to_save.columns:
                        if c in row:
                            df_to_save.loc[mask, c] = row[c]
                else:
                    df_to_save = pd.concat([df_to_save, pd.DataFrame([row])], ignore_index=True)
                    
            salvar_catalogo(df_to_save)
            st.session_state['reset_counter_padconf'] += 1
            st.success("✅ Catálogo atualizado com sucesso!")
            st.rerun()

        st.divider()
        col_sync, _ = st.columns([3, 7])

        with col_sync:
            if st.button("📥 Puxar Nomes do ERP", use_container_width=True):
                try:
                    conn_pg = st.connection("banco_erp", type="sql")
                    cods = tuple(edited_cat["Código"].tolist())

                    if len(cods) == 1:
                        cods_str = f"({cods[0]})"
                    else:
                        cods_str = str(cods)

                    query_nomes = f"SELECT cadp_codigo, cadp_descricao FROM cadprod WHERE cadp_codigo IN {cods_str}"
                    df_nomes = conn_pg.query(query_nomes, ttl=0)

                    for _, row in df_nomes.iterrows():
                        mask = edited_cat["Código"] == row["cadp_codigo"]
                        edited_cat.loc[mask, "Descrição Oficial"] = row["cadp_descricao"]

                    df_to_save = carregar_catalogo_padconf().drop(columns=["Descrição"], errors="ignore")
                    for _, row in edited_cat.iterrows():
                        mask = (df_to_save["Código"] == row["Código"]) & (df_to_save["Fornecedor"] == row["Fornecedor"])
                        if mask.any():
                            df_to_save.loc[mask, "Descrição Oficial"] = row["Descrição Oficial"]
                            
                    salvar_catalogo(df_to_save)
                    st.success("✅ Nomes Oficiais sincronizados com sucesso!")
                    st.rerun()
                except Exception as e:
                    if "No database configured" in str(e) or "missing" in str(e).lower():
                         st.error("⚠️ Aviso: As credenciais de acesso ao PostgreSQL não foram encontradas no painel Secrets.")
                    else:
                         st.error(f"⚠️ Erro ao buscar nomes no banco ERP: {e}")
